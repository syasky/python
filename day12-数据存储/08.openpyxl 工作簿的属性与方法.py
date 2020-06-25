
#
from openpyxl import Workbook    #创建新工作簿的
from openpyxl import load_workbook   #打开已有工作簿
#注意： load_workbook 打不开xls文件

wb=load_workbook('某公司贸易数据.xlsx')
'''
wb 工作簿对象
属性：
wb.sheetnames    :  所有工作表名称 组成的列表
wb.worksheets    ： 所有工作表 组成的列
wb.active        :  默认的工作表 （第一张）

方法：
wb.get_sheet_names()  : 同  wb.sheetnames  
wb.get_active_sheet() : 同  wb.active 
wb.get_sheet_by_name( sheetName ) : 根据表的名称 获取 指定 sheet
wb.create_sheet(sheetName ,index ) :  创建并返回 sheet ，名叫sheetName，在index 位置
wb.save(fileName)

'''
# print(wb.sheetnames)
#创建一个表单
sheet=wb.create_sheet('我的地盘',0)
sheet.append([1,2,3])
sheet.append([4,5,6])
wb.save('某公司贸易数据.xlsx')

