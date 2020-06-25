from openpyxl import load_workbook

from openpyxl.styles import colors
from openpyxl.styles import Font,Color
#打开工作簿
wb =load_workbook('../day12-数据存储/某公司贸易数据.xlsx')
#获取一个表格对象
# 更推荐 字典的方法  读取某个表
sheet=wb['产品类别']

'''  工作表的对象的属性 与 方法
sheet.rows            : 行数对象
sheet.columns         : 列数对象
sheet.max_row         : 有效的最大行数
sheet.min_row         : 有效的最小行数
sheet.max_column      : 有效的最大列数
sheet.min_column      : 有效的最小列数
sheet.values           :所有单元格的值组成的对象，可以理解为二维列表，每行是一个元组
sheet.title           :表名称

方法：
sheet.cell(row=1,column=1)    获取单元格，从1开始计数
'''
print(list(sheet.rows))
print(sheet.max_row)
print(sheet.min_row)

#
print(sheet.values)
print(list(sheet.values))
print(sheet.cell(1,1))
print(sheet.cell(1,1).value)

#注意  还可以通过  a1  b2  的形式，以字典的形式直接读取单元格，A1，a1都行
print(sheet['b2'].value)

# 写入呢?我们已经可以append整行写入了单个呢?
# sheet[ 'B2']. value = xur  进行赋值即可
#如将c4单元格变成炸弹

font=Font(size=30,color='cf79fe')
sheet['c4'].value='炸弹'
sheet['c4'].font=font

#可以按列写入
for i in range(1,11):
    #sheet[f'+str(i)'].value= i*10
    sheet[f'F{i}'].value=i*10

wb.save('某公司新-----1.xlsx')